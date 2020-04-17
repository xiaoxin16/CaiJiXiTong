#! /usr/bin/python
# -*- coding: utf8 -*-

import re, time, ipdb, math, os, socket, datetime
from multiprocessing import Pool
from urllib.parse import urlparse
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, UnexpectedAlertPresentException
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from PIL import Image, ImageDraw, ImageFont
import platform


# import logging
#
# module_logger = logging.getLogger("Main.sub")

# 1.
def isIP(str):
    p = re.compile('^((25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(25[0-5]|2[0-4]\d|[01]?\d\d?)$')
    if p.match(str):
        return True
    else:
        return False


# 2. url formalization 标准化
def get_url_normalize_single(url):
    url_new = url
    default_scheme = "http"
    if urlparse(url).scheme == '':
        url_new = default_scheme + "://" + url
    elif urlparse(url).scheme == 'https':
        url_new = urlparse(url).scheme + "://" + urlparse(url).hostname
    if urlparse(url_new).hostname is None:
        url_new = "异常"
    elif "." not in urlparse(url_new).hostname:
        url_new = "异常"
    elif urlparse(url_new).hostname[0] is ".":
        url_new = "异常"
    elif len(urlparse(url_new).hostname) < 4:
        url_new = "异常"
    else:
        {}
    url_site = url_new
    return url_site


# 2. 非空IP 列表
def update_task_excel(da, conf_data, add_head):
    wb = load_workbook(conf_data["dst"] + "/" + conf_data["fn"])
    ws = wb.active
    alig_s = Alignment(horizontal='left', vertical='center')
    god_sheet_name = '核查结果'
    # bad_sheet_name = '核查结果_异常'
    if god_sheet_name not in wb.sheetnames:
        ws_1 = wb.create_sheet(god_sheet_name, 0)
    else:
        ws_1 = wb[god_sheet_name]
    index_1 = 1
    if add_head == 1:
        for i in range(len(conf_data["title"])):
            ws_1.cell(row=index_1, column=i + 1).value = conf_data["title"][i]
            ws_1.cell(row=index_1, column=i + 1).alignment = alig_s
        index_1 = 2
    for i in sorted(da.keys()):
        for ii in range(1, len(da[i]) + 1):
            ws_1.cell(row=i+index_1-1, column=ii).value = da[i][ii - 1]
            ws_1.cell(row=i+index_1-1, column=ii).alignment = alig_s
        # print(da[i])
        # if da[i][4].strip() == "" or da[i][4].strip() == "NULL":
        #     for ii in range(1, len(da[i]) + 1):
        #         ws_2.cell(row=index_2, column=ii).value = da[i][ii - 1]
        #         ws_2.cell(row=index_2, column=ii).alignment = alig_s
        #     index_2 = index_2 + 1
        # else:
        #     for ii in range(1, len(da[i]) + 1):
        #         ws_1.cell(row=index_1, column=ii).value = da[i][ii - 1]
        #         ws_1.cell(row=index_1, column=ii).alignment = alig_s
        #     index_1 = index_1 + 1
    wb.save(conf_data["dst"] + "/" + conf_data["fn"])


# 3. multiprocess_fun
def multiprocess_fun(d_a, task_kind, conf_data):
    poll_num = conf_data["poll"]
    pool = Pool(poll_num)
    key_dic = {}
    value_dic = {}
    dict_emp = {}
    # 分组
    for i in range(poll_num):
        key_dic[i] = []
        value_dic[i] = []
    for key, value in d_a.items():
        if task_kind == 2:
            if value[4] == "" or value[4] == "NULL":
                value.append("NULL")
                value.append("NULL")
                value.append("NULL")
                dict_emp[key] = value

                screen_shot_file_name = conf_data["screenshot"] + "/" + str(value[0]) + "_" + str(value[3]) + ".png"
                img = Image.new('RGB', (1366, 768), (255, 255, 255))
                img.save(screen_shot_file_name)
                img = Image.open(screen_shot_file_name)
                draw = ImageDraw.Draw(img)
                if platform.system() == 'Windows':
                    font_info = "C:\\Windows\\Fonts\\SIMLI.TTF"
                elif platform.system() == 'Linux':
                    font_info = "C:\\Windows\\Fonts\\SIMLI.TTF"
                else:
                    print('其他')
                ttfont = ImageFont.truetype(font=font_info, size=80)
                draw.text((550, 330), u"解析失败", fill="#0000ff", font=ttfont)
                img.save(screen_shot_file_name)
                # fpt = open(screen_shot_file_name, 'w')
                # fpt.write("解析失败".encode("utf-8"))
                continue
        for i in range(poll_num):
            if int(key % poll_num) == i:
                key_dic[i].append(key)
                value_dic[i].append(value)
                continue
    new_dict = {}
    res_dict = []
    for i in range(poll_num):
        new_dict[i] = dict(zip(key_dic[i], value_dic[i]))
        # module_logger.info("分组:%s" % (new_dict[i]))
        if task_kind == 1:
            res_dict.append(pool.apply_async(dns_process, (new_dict[i], conf_data, i)))
        elif task_kind == 2:
            res_dict.append(pool.apply_async(get_title_by_selenium, (new_dict[i], conf_data, i)))
        elif task_kind == 3:
            res_dict.append(pool.apply_async(get_alexa_rank_by_link114, (new_dict[i], conf_data, i)))
        elif task_kind == 4:
            res_dict.append(pool.apply_async(get_alexa_rank_by_link114_multi, (new_dict[i], conf_data, i)))
    pool.close()
    pool.join()

    res_dict_N = {}
    res_dict_JS = {}
    for res in res_dict:
        res_dict_N.update(res.get())
    if task_kind == 1:
        update_task_excel(res_dict_N, conf_data, 0)
        return res_dict_N
    elif task_kind == 2:
        res_dict_N.update(dict_emp)
        for key, value in res_dict_N.items():
            if value[7] == "javascript":
                res_dict_JS[key] = value
        update_task_excel(res_dict_N, conf_data, 0)
        return res_dict_JS
        # if len(res_dict_JS) > 0:
        #     return res_dict_JS
        # else:
        #     return res_dict_N
    elif task_kind == 3:
        update_task_excel(res_dict_N, conf_data, 1)
        return res_dict_N
    elif task_kind == 4:
        update_task_excel(res_dict_N, conf_data, 1)
        return res_dict_N


# 4. [3, 'http://01014688.comxxx', 'http://01014688.comxxx', '01014688.comxxx', '', '', '', '']
def dns_process(d_a, conf_data, i):
    import logging
    logger = logging.getLogger("DNS")
    logger.setLevel(level=logging.INFO)
    handler = logging.FileHandler("%s/SELENIUM-%d-log.txt" % (conf_data["log"], i), encoding='utf-8')
    handler.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)

    # console set
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)

    logger.addHandler(handler)
    # logger.addHandler(console)

    start = datetime.datetime.now()

    for key, value in d_a.items():
        norm_url = get_url_normalize_single(value[1])
        value.append(norm_url)
        domain = urlparse(norm_url).hostname
        myaddr = []
        if isIP(domain):
            value.append("")
            value.append(domain)
            myaddr.append(domain)
            # logger.info("[%s\t%s\t%s]" % (value[0], domain, domain))
        else:
            value.append(domain)
            try:
                A = socket.gethostbyname(domain)
                myaddr.append(A)
                # logger.info("[%s\t%s\t%s]" % (value[0], domain, myaddr[0]))
            except socket.error:
                myaddr.append("")
                # logger.info("[%s\t%s\t解析失败]" % (value[0], domain))
        time.sleep(0.1)
        if myaddr[0].strip() == "":
            value.append("NULL")
            value.append("NULL")
            value.append("NULL")
            msgstr = ",".join(map(str, value))
            logger.info("[%s]" % msgstr)
        else:
            value.append(myaddr[0])
            dbpath = conf_data["conf"] + "/ipipfree.ipdb"
            db = ipdb.City(dbpath)
            city_str = db.find(myaddr[0], "CN")
            jing_nei = "中国,香港,澳门,台湾"
            if city_str[0] in jing_nei:
                if city_str[1] in jing_nei:
                    value.append("境外")
                else:
                    value.append("境内")
            else:
                value.append("境外")
            value.append(city_str[0] + "·" + city_str[1])
    end = datetime.datetime.now()
    # logger.info("%d" % (end-start).seconds)
    return d_a


# 5.
def get_title_by_selenium(d_a, conf_data, i):
    import logging
    logger = logging.getLogger("SELENIUM")
    logger.setLevel(level=logging.INFO)
    handler = logging.FileHandler("%s/SELENIUM-%d-log.txt" % (conf_data["log"], i), encoding='utf-8')
    handler.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)

    # console set
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)

    logger.addHandler(handler)
    # logger.addHandler(console)

    start = datetime.datetime.now()
    chrome_options = Options()
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--window-size=1366,768')
    # chrome_options.add_argument('--headless')
    # if conf_data["javascript"] == "close":
    #     pref_sets ={
    #         'profile.default_content_setting_values': {
    #             'javascript': 2
    #         }
    #     }
    #     chrome_options.add_experimental_option('prefs', pref_sets)
    chrome_options.add_argument("enable-automation")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-browser-side-navigation")
    chrome_options.add_argument("--disable-gpu")

    browser = webdriver.Chrome(executable_path=conf_data["conf"] + "/chromedriver.exe", options=chrome_options)
    timeout_s = 10
    browser.implicitly_wait(timeout_s)
    screen_shot_dir = conf_data["screenshot"] + "/"
    page_source_dir = conf_data["pagesource"] + "/"
    col_l = conf_data["col"]
    yum = len(d_a) % col_l
    counts = math.ceil(len(d_a) / col_l)
    key_list = list(d_a.keys())
    value_list = list(d_a.values())
    col_add = col_l
    logger.info("key_list:%s" % key_list)
    logger.info("value_list:%s" % value_list)
    for i in range(counts):
        if (i == (counts - 1)) and (yum > 0):
            col_add = yum
        # print("单个进程轮询:", len(d_a), yum, counts)
        for j in range(col_add):
            index = value_list[i * col_l + j][0]
            url = value_list[i * col_l + j][2]
            # logger.info("%s, %s" % (index, url))
            # 开始请求
            try:
                js = "window.open(\"" + url + "\");"
                browser.execute_script(js)
                logger.info("%s JS execute OK" % url)
            except {socket.timeout, TimeoutException}:
                logger.info("%s socket超时:" % url)
                # browser.execute_script("window.stop();")
                url_refer = "超时"
        time.sleep(10)
        for j in range(col_add):
            index = value_list[i * col_l + j][0]
            url = value_list[i * col_l + j][2]
            all_win = browser.window_handles
            browser.switch_to.window(all_win[-1])
            domain = urlparse(url).hostname
            logger.info("切换\t%s" % url)
            try:
                b_title = browser.title
                if b_title == "":
                    b_title = "补" + url
                # print(index, "\t", url, "\t", b_title)
                url_con = browser.current_url.rstrip('/')
                url_new = url.rstrip('/')
                if url_new in url_con:
                    url_refer = "NULL"
                else:
                    url_refer = browser.current_url
                browser.get_screenshot_as_file(screen_shot_dir + str(index) + "_" + str(domain) + ".png")

                fb = open(page_source_dir + str(index) + "_" + str(domain) + ".html", 'wb')
                fb.write(browser.page_source.encode("utf-8", "ignore"))
                fb.close()

                if "浏览器需要支持JavaScript" in browser.page_source:
                    b_title = "javascript"
                    logger.info("%s\t%s\t%s, 正常-截图成功-浏览器需要支持JavaScript" % (index, url, b_title))
                elif "load javascript" in browser.page_source:
                    b_title = "javascript"
                    logger.info("%s\t%s\t%s, 正常-截图成功-浏览器需要支持JavaScript" % (index, url, b_title))
                else:
                    logger.info("%s\t%s\t%s, 正常-截图成功" % (index, url, b_title))
            except UnexpectedAlertPresentException:
                browser.switch_to.alert.accept()
                b_title = browser.title
                if b_title == "":
                    b_title = "补" + url
                # print(index, "\t", url, "\t", b_title)
                url_con = browser.current_url.rstrip('/')
                if domain == urlparse(url_con).hostname:
                    url_refer = "NULL"
                else:
                    url_refer = browser.current_url
                fb = open(page_source_dir + str(index) + "_" + str(domain) + ".html", 'wb')
                fb.write(browser.page_source.encode("utf-8", "ignore"))
                fb.close()

                browser.get_screenshot_as_file(screen_shot_dir + str(index) + "_" + str(domain) + ".png")
                logger.info("%s\t%s\t%s, 弹窗-截图成功" % (index, url, b_title))

            except TimeoutException:
                url_refer = "超时"
                b_title = "超时或者无法访问"
                url_con = "NULL"
                try:
                    browser.get_screenshot_as_file(screen_shot_dir + str(index) + "_" + str(domain) + ".png")
                    b_title = "补" + url

                    fb = open(page_source_dir + str(index) + "_" + str(domain) + ".html", 'wb')
                    fb.write(browser.page_source.encode("utf-8", "ignore"))
                    fb.close()

                    logger.info("%s\t%s\t%s, 超时-截图成功" % (index, url, b_title))
                except BaseException as msg:
                    screen_shot_file_name = screen_shot_dir + str(index) + "_" + str(domain) + ".png"
                    img = Image.new('RGB', (1366, 768), (255, 255, 255))
                    img.save(screen_shot_file_name)
                    img = Image.open(screen_shot_file_name)
                    draw = ImageDraw.Draw(img)
                    if platform.system() == 'Windows':
                        font_info = conf_data["conf"] + "/SIMLI.TTF"
                    elif platform.system() == 'Linux':
                        font_info = conf_data["conf"] + "/SIMLI.TTF"
                    else:
                        print('其他')
                    ttfont = ImageFont.truetype(font=font_info, size=80)
                    draw.text((550, 330), u"超 时", fill="#0000ff", font=ttfont)
                    img.save(screen_shot_file_name)

                    fb = open(page_source_dir + str(index) + "_" + str(domain) + ".html", 'wb')
                    fb.write(browser.page_source.encode("utf-8", "ignore"))
                    fb.close()

                    logger.info("%s, 超时-截图失败, %s" % (url, str(msg)))
                except TimeoutException:
                    print("未知超时")
                    logger.info("%s, 超时-截图失败, 未知原因" % url)
            browser.close()
            browser.switch_to.window(all_win[0])
            value_list[i * col_l + j].append(b_title)
            value_list[i * col_l + j].append(url_refer)
            value_list[i * col_l + j].append(url_con)
            msgstr = ",".join(map(str, value_list[i * col_l + j]))
            logger.info("[%s]" % msgstr)
    browser.close()
    browser.quit()
    end = datetime.datetime.now()
    logger.info("耗时:%d秒" % (end - start).seconds)
    return d_a


# 6. link114 get Alexa one by one
def get_alexa_rank_by_link114(da, conf_data, i):
    url_114 = "http://www.link114.cn/alexa/"
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    browser = webdriver.Chrome(executable_path=conf_data["conf"] + "/chromedriver.exe", options=chrome_options)
    timeout_s = 20
    browser.implicitly_wait(timeout_s)
    # 开始请求
    # js = "window.open(\"" + url_114 + "\");"
    # browser.execute_script(js)
    browser.get(url_114)
    for key, value in da.items():
        url = value[3]
        if url is not None:
            browser.find_element_by_id("ip_websites").clear()
            browser.find_element_by_id("ip_websites").send_keys(url)
            browser.find_element_by_id("tj").click()
            trlist = browser.find_elements_by_tag_name("tr")
            for tr in trlist:
                tdlist = tr.find_elements_by_tag_name("td")
                if len(tdlist) > 0:
                    trid_1 = tr.find_elements_by_tag_name("td")[1].get_attribute("value")
                    trid_alexa = tr.find_elements_by_tag_name("td")[2].text
                    trid_alexa = trid_alexa.replace("Alexa:", "")
                    value.append(trid_1)
                    value.append(trid_alexa)
        time.sleep(0.5)
        # print(value)
    browser.close()
    browser.quit()
    return da


# 7. link114 get Alexa column
def get_alexa_rank_by_link114_multi(d_a, conf_data, i):
    import logging
    logger = logging.getLogger("Alexa")
    logger.setLevel(level=logging.INFO)
    handler = logging.FileHandler("%s/SELENIUM-%d-log.txt" % (conf_data["log"] , i), encoding='utf-8')
    handler.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)

    # console set
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)

    logger.addHandler(handler)
    # logger.addHandler(console)

    start = datetime.datetime.now()

    chrome_options = Options()
    chrome_options.add_argument('--window-size=1366,768')
    chrome_options.add_argument('--headless')
    url_114 = "http://www.link114.cn/"
    browser = webdriver.Chrome(executable_path=conf_data["conf"] + "/chromedriver.exe", options=chrome_options)
    timeout_s = 5
    browser.implicitly_wait(timeout_s)
    # 开始请求
    browser.get(url_114)
    # unselect 1
    str1 = "//*[@id='chk_baidu_qz_zz']"
    check_conditions = browser.find_element_by_xpath(str1)
    browser.execute_script("$(arguments[0]).click();", check_conditions)
    browser.execute_script("$(arguments[0]).attr('checked',false);", check_conditions)
    # unselect 2
    str2 = "//*[@id='chk_baidu_qz_ai']"
    check_conditions = browser.find_element_by_xpath(str2)
    browser.execute_script("$(arguments[0]).click();", check_conditions)
    browser.execute_script("$(arguments[0]).attr('checked',false);", check_conditions)
    # unselect 3
    str3 = "//*[@id='chk_baidu_sl']"
    check_conditions = browser.find_element_by_xpath(str3)
    browser.execute_script("$(arguments[0]).click();", check_conditions)
    browser.execute_script("$(arguments[0]).attr('checked',false);", check_conditions)
    # unselect 4
    str4 = "//*[@id='chk_so360_qz_zz']"
    check_conditions = browser.find_element_by_xpath(str4)
    browser.execute_script("$(arguments[0]).click();", check_conditions)
    browser.execute_script("$(arguments[0]).attr('checked',false);", check_conditions)

    # select 5
    # str5 = "//*[@id='chk_title']"
    # check_conditions = browser.find_element_by_xpath(str5)
    # browser.execute_script("$(arguments[0]).click();", check_conditions)
    # browser.execute_script("$(arguments[0]).attr('checked',true);", check_conditions)
    # select 6
    str6 = "//*[@id='chk_alexa']"
    check_conditions = browser.find_element_by_xpath(str6)
    browser.execute_script("$(arguments[0]).click();", check_conditions)
    browser.execute_script("$(arguments[0]).attr('checked',true);", check_conditions)
    # select 7
    # str7 = "//*[@id='chk_ip']"
    # check_conditions = browser.find_element_by_xpath(str7)
    # browser.execute_script("$(arguments[0]).click();", check_conditions)
    # browser.execute_script("$(arguments[0]).attr('checked',true);", check_conditions)

    col_l = conf_data["col"]
    yum = len(d_a) % col_l
    counts = math.ceil(len(d_a) / col_l)
    key_list = list(d_a.keys())
    value_list = list(d_a.values())
    col_add = col_l
    # logger.info("key_list:%s" % key_list)
    # logger.info("value_list:%s" % value_list)
    for i in range(counts):
        if (i == (counts - 1)) and (yum > 0):
            col_add = yum
        domain_set = []
        for j in range(col_add):
            index = value_list[i * col_l + j][0]
            domain = value_list[i * col_l + j][3]
            domain_set.append(domain)
        domain_str = ','.join(domain_set)
        # logger.info(domain_str)
        browser.find_element_by_id("ip_websites").clear()
        browser.find_element_by_id("ip_websites").send_keys(domain_str)
        browser.find_element_by_id("tj").click()
        time.sleep(10)
        trlist = browser.find_elements_by_tag_name("tr")
        data_dict = {}
        index = 0
        for tr in trlist:
            # 获取tr中的所有td
            tdlist = tr.find_elements_by_tag_name("td")
            data_dict[index] = []
            if len(tdlist) > 0:
                # 获取td[0]的文本
                text_1 = tr.find_elements_by_tag_name("td")[0].text
                text_1 = text_1.replace(".", "")
                # trid_1 = tr.get_attribute("id")
                trid_1 = tr.find_elements_by_tag_name("td")[1].get_attribute("value")
                trid_alexa = tr.find_elements_by_tag_name("td")[2].text
                trid_alexa = trid_alexa.replace("Alexa:", "")
                data_dict[index].append(text_1)
                data_dict[index].append(trid_1)
                data_dict[index].append(trid_alexa)
                index = index + 1
        alexa_set = data_dict.values()
        for d in alexa_set:
            for key, value in d_a.items():
                if d[1] == value[3]:
                    value.append(d[1])
                    value.append(d[2])
                    msgstr = ",".join(map(str, value))
                    logger.info("[%s]" % msgstr)
                    break
        time.sleep(timeout_s)

    browser.close()
    browser.quit()
    end = datetime.datetime.now()

    logger.info("耗时:%d秒" % (end - start).seconds)
    return d_a
