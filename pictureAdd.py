#!/usr/bin/python
# -*- coding:utf-8 -*-

import datetime, os


# 1. 读取配置文件
def read_conf(conf_path):
    import json
    with open(conf_path, encoding='utf-8-sig') as f:
        data = json.load(f)
    # for key, value in data.items():
    #     print("配置文件：", key, ":", value)
    return data


# 2. 读取待处理文件-excel
# 2.1 选择文件, 并复制到目录文件
def select_file(fp, dst_fp):
    import os, shutil
    files = os.listdir(fp)
    files_set = []
    print("文件列表：")
    for file in files:
        if os.path.splitext(file)[1] == ".xlsx":
            files_set.append(file)
            print(len(files_set), ':', file)
    index = input("请输入对应文件的序号:")
    file_name = files_set[int(index)-1]
    dst_file = "核查结果_" + file_name
    if not os.path.exists(dst_fp + "/" + dst_file):
        shutil.copy(fp + "/" + file_name, dst_fp + "/" + dst_file)
    return dst_file


# 2.2 读取文件内容
def get_excel_data(fp, fn, width, isprocess):
    from openpyxl import Workbook
    from openpyxl import load_workbook

    wb = load_workbook(fp + "/" + fn)
    ws = wb.worksheets[0]
    data_arry = {}
    for r in range(ws.max_row):
        if isprocess == 1:
            if (ws.cell(row=r+1, column=width+1).value is None) and (ws.cell(row=r+1, column=width+2).value is None):
                data_arry[r + 1] = []
                for i in range(1, width+1):
                    data_arry[r + 1].append(ws.cell(row=r+1, column=i).value)
        else:
            data_arry[r + 1] = []
            for i in range(1, width + 1):
                data_arry[r + 1].append(ws.cell(row=r + 1, column=i).value)
    return data_arry


def add_is_db_based_screen(d_a, fpath):
    import os
    index_set = []
    for root, dirs, files in os.walk(fpath):
        for file in files:
            index = file.split('_')[0]
            index_set.append(int(index))
    print(len(index_set))
    for key, value in d_a.items():
        # print(key, value)
        if key == 1:
            value.append("是否有害")
            continue
        if value[0] in index_set:
            value.append("是")
            index_set.remove(value[0])
        else:
            value.append("待定")
    return d_a


# final: write file
def write_info_excel(f_p, fn, da, start_index):
    from openpyxl import Workbook
    from openpyxl import load_workbook
    wb = load_workbook(f_p + "/" + fn)
    ws = wb.worksheets[0]
    for key, value in da.items():
        for i in range(start_index, len(value)+1):
            ws.cell(row=key, column=i).value = value[i-1]
    wb.save(f_p + "/" + fn)


def main():
    conf_fp = "../workdata/conf/config.json"
    start = datetime.datetime.now()
    print("1. 加载配置文件...")
    conf_data = read_conf(conf_fp)
    print("2. 加载任务文件，复制到输出目录 ...")
    file_name = select_file(conf_data["src"], conf_data["dst"])
    conf_data["fn"] = file_name
    da = get_excel_data(conf_data["dst"], file_name, 12, 0)
    add_is_db_based_screen(da, conf_data["dst"] + "/截图")
    write_info_excel(conf_data["dst"], file_name, da, 12)
    end = datetime.datetime.now()
    print("final is in ", (end - start).seconds)
    os.system("pause")


if __name__ == "__main__":
    main()