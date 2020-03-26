#!/usr/bin/python
# -*- coding:utf-8 -*-

import datetime, os


# 1. 读取配置文件
def read_conf(conf_path):
    import json
    with open(conf_path, encoding='utf-8-sig') as f:
        data = json.load(f)
    return data


# 2. 读取待处理文件-excel
# 2.1 选择文件, 并复制到目录文件
def select_file(fp):
    import os, shutil
    files = os.listdir(fp)
    files_set = []
    print("文件列表：")
    for file in files:
        if os.path.splitext(file)[1] == ".xlsx":
            files_set.append(file)
            print(len(files_set), ':', file)
    index = input("请输入对应excel文件的序号:")
    file_name = files_set[int(index)-1]
    return file_name


# 2.2 读取文件内容
def get_excel_data(fn, width, isprocess):
    from openpyxl import Workbook
    from openpyxl import load_workbook

    wb = load_workbook(fn)
    ws = wb.worksheets[0]
    data_arry = {}
    for r in range(ws.max_row):
        if isprocess == 1:
            if (ws.cell(row=r+1, column=width+1).value is None) and (ws.cell(row=r+1, column=width+2).value is None):
                data_arry[r + 1] = []
                for i in range(1, width+1):
                    data_arry[r + 1].append(ws.cell(row=r+1, column=i).value)
        else:
            data_arry[r + 1] = []
            for i in range(1, ws.max_column + 1):
                data_arry[r + 1].append(ws.cell(row=r + 1, column=i).value)
    return data_arry


def add_is_db_based_screen(d_a, fpath):
    import os
    index_set = []
    for root, dirs, files in os.walk(fpath):
        for file in files:
            index = file.split('_')[0]
            index_set.append(int(index))
    # print(len(index_set), index_set)
    for key, value in d_a.items():
        # print(key, value)
        if key == 1:
            value.append("是否有害")
            continue
        if int(value[0]) in index_set:
            value.append("是")
            index_set.remove(value[0])
        else:
            value.append("否")
    return d_a


# final: write file
def write_info_excel(fn, da):
    from openpyxl import Workbook
    from openpyxl import load_workbook
    wb = load_workbook(fn)
    ws = wb.worksheets[0]
    for key, value in da.items():
        for i in range(1, len(value)+1):
            ws.cell(row=key, column=i).value = value[i-1]
    wb.save("新_" + fn)


def main():
    start = datetime.datetime.now()
    file_name = select_file("./")
    da = get_excel_data(file_name, 12, 0)
    add_is_db_based_screen(da, "./截图")
    write_info_excel(file_name, da)
    end = datetime.datetime.now()
    print("运行结束，耗时%d秒 " % (end - start).seconds)
    os.system("pause")


if __name__ == "__main__":
    main()