#!/usr/bin/python
# -*- coding:utf-8 -*-

import datetime, os
import logging
from multiprocessing import freeze_support
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment


# 1. 读取配置文件
def read_conf(conf_path):
    import json
    with open(conf_path, encoding='utf-8-sig') as f:
        data = json.load(f)
    return data


# 2. 读取待处理文件-excel
# 2.1 选择文件, 并复制到目录文件
def select_file(conf):
    import shutil, re
    files = os.listdir(conf["log"])
    files_set = []
    data = {}
    print("文件列表：")
    for file in files:
        if os.path.splitext(file)[1] == ".txt":
            files_set.append(file)
            print('[', len(files_set), ']:', file)
            if "main" in file:
                continue
            with open(conf["log"] + "/" + file, encoding='utf-8') as f:
                line = f.readline()
                while line:
                    # print(line)
                    line.rstrip("\n")
                    if "DNS - INFO - [" in line:
                        newl = re.findall(r'[[](.*)[]]', line)
                        strs = newl[0].split(',')
                        data[int(strs[0])] = strs
                    elif "SELENIUM - INFO - [" in line:
                        newl = re.findall(r'[[](.*)[]]', line)
                        strs = newl[0].split(',')
                        data[int(strs[0])] = strs
                    line = f.readline()
    for key, value in data.items():
        print(key, len(value), value)
    return data


# 2.2 读取文件内容
def get_excel_data(fp, fn, width, isprocess):

    wb = load_workbook(fp + "/" + fn)
    ws = wb.worksheets[0]
    data_arry = {}
    for r in range(ws.max_row):
        if isprocess == 1:
            if (ws.cell(row=r+1, column=width+1).value is None) and (ws.cell(row=r+1, column=width+2).value is None):
                data_arry[r + 1] = []
                for i in range(1, width+1):
                    data_arry[r + 1].append(ws.cell(row=r+1, column=i).value)
        else:
            data_arry[r + 1] = []
            for i in range(1, width + 1):
                data_arry[r + 1].append(ws.cell(row=r + 1, column=i).value)
    return data_arry


# 2. 非空IP 列表
def write_task_excel(da, conf_data, add_head):
    wb = Workbook()
    ws = wb.active
    alig_s = Alignment(horizontal='left', vertical='center')
    ws_1 = wb.create_sheet("核查结果", 0)
    index_1 = 1
    if add_head == 1:
        for i in range(len(conf_data["title"])):
            ws_1.cell(row=1, column=i + 1).value = conf_data["title"][i]
            ws_1.cell(row=1, column=i + 1).alignment = alig_s
        index_1 = 2
    for i in sorted(da):
        for ii in range(1, len(da[i]) + 1):
            if ii == 1:
                ss = int(da[i][ii - 1])
            else:
                ss = da[i][ii - 1]
            ws_1.cell(row=index_1, column=ii).value = ss
            ws_1.cell(row=index_1, column=ii).alignment = alig_s
        index_1 = index_1 + 1

    wb.save(conf_data["dst"] + "/日志核查结果.xlsx")


def main():
    # begin
    conf_fp = "./data/conf/config.json"
    conf_data = read_conf(conf_fp)
    da = select_file(conf_data)
    write_task_excel(da, conf_data, 1)

    # os.system("pause")


if __name__ == "__main__":
    freeze_support()
    main()
