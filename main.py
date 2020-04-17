#!/usr/bin/python
# -*- coding:utf-8 -*-

import datetime, io, os, sys
import logging
from multiprocessing import freeze_support
import WebInfo


# 1. 读取配置文件
def read_conf(conf_path):
    import json
    with open(conf_path, encoding='utf-8-sig') as f:
        data = json.load(f)
    # for key, value in data.items():
    #     print("配置文件：", key, ":", value)
    return data


# 2. 读取待处理文件-excel
# 2.1 选择文件, 并复制到目录文件
def select_file(fp, dst_fp):
    import shutil
    files = os.listdir(fp)
    files_set = []
    print("文件列表：")
    for file in files:
        if os.path.splitext(file)[1] == ".xlsx":
            files_set.append(file)
            print('[', len(files_set), ']:', file)
    index = input("请输入对应文件的序号:")
    file_name = files_set[int(index)-1]
    dst_file = file_name
    if not os.path.exists(dst_fp + "/" + os.path.splitext(dst_file)[0]):
        os.mkdir(dst_fp + "/" + os.path.splitext(dst_file)[0])
        shutil.copy(fp + "/" + file_name, dst_fp + "/" + os.path.splitext(dst_file)[0] + "/" + dst_file)
    return dst_file


# 2.2 读取文件内容
def get_excel_data(fp, fn, width, isprocess):
    from openpyxl import Workbook
    from openpyxl import load_workbook

    wb = load_workbook(fp + "/" + fn)
    ws = wb.worksheets[0]
    data_arry = {}
    for r in range(ws.max_row):
        if isprocess == 1:
            if (ws.cell(row=r+1, column=width+1).value is None) and (ws.cell(row=r+1, column=width+2).value is None):
                data_arry[r + 1] = []
                for i in range(1, width+1):
                    data_arry[r + 1].append(ws.cell(row=r+1, column=i).value)
        else:
            data_arry[r + 1] = []
            for i in range(1, width + 1):
                data_arry[r + 1].append(ws.cell(row=r + 1, column=i).value)
    return data_arry


def main():
    # begin
    # sys.stdout.reconfigure(encoding='utf-8')
    conf_fp = "./data/conf/config.json"
    conf_data = read_conf(conf_fp)
    if not os.path.exists(conf_data["dst"]):
        os.mkdir(conf_data["dst"])

    file_name = select_file(conf_data["src"], conf_data["dst"])
    conf_data["fn"] = file_name
    conf_data["dst"] = conf_data["dst"] + "/" + os.path.splitext(conf_data["fn"])[0]

    conf_data["log"] = conf_data["dst"] + "/" + conf_data["log"]
    if not os.path.exists(conf_data["log"]):
        os.mkdir(conf_data["log"])

    conf_data["screenshot"] = conf_data["dst"] + "/" + "/" + conf_data["screenshot"]
    if not os.path.exists(conf_data["screenshot"]):
        os.mkdir(conf_data["screenshot"])

    conf_data["pagesource"] = conf_data["dst"] + "/" + "/" + conf_data["pagesource"]
    if not os.path.exists(conf_data["pagesource"]):
        os.mkdir(conf_data["pagesource"])

    # log set
    logger = logging.getLogger("Main")
    logger.setLevel(level=logging.INFO)

    handler = logging.FileHandler("%s/main_log.txt" % (conf_data["log"]), encoding='utf-8')
    handler.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)

    # console set
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)

    logger.addHandler(handler)
    logger.addHandler(console)

    start = datetime.datetime.now()
    logger.info("1. 加载配置文件")

    logger.info("2. 加载任务文件，复制到输出目录")

    da = get_excel_data(conf_data["dst"], file_name, 2, 0)
    logger.info("3. 加载excel完毕，开始DNS, len=%d" % (len(da)))
    da = WebInfo.multiprocess_fun(da, 1, conf_data)
    da = get_excel_data(conf_data["dst"], file_name, 7, 0)
    logger.info("4. 加载excel完毕，开始Chrome, len=%d" % (len(da)))
    da = WebInfo.multiprocess_fun(da, 2, conf_data)
    if len(da) > 0:
        conf_data["javascript"] = "open"
        da = WebInfo.multiprocess_fun(da, 2, conf_data)
    da = get_excel_data(conf_data["dst"], file_name, 10, 0)
    logger.info("5. 加载excel完毕，开始查询排名, len=%d" % (len(da)))
    da = WebInfo.multiprocess_fun(da, 4, conf_data)
    end = datetime.datetime.now()
    logger.info("6. final is in %d" % (end - start).seconds)
    # os.system("pause")


if __name__ == "__main__":
    freeze_support()
    main()
